"""
Excel export for the Performance page.

Builds the same figures the page shows, honouring the selected date window and operator, so a
downloaded workbook can be reconciled against the screen. Sheets mirror the page's blocks.
"""
import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import app.services.agenda_service as agenda_service
import app.services.leads_service as leads_service
import app.services.negocios_service as negocios_service

_HEADER_FILL = PatternFill("solid", fgColor="1F2937")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
_TITLE_FONT = Font(bold=True, size=12)
_LABEL_FONT = Font(bold=True, size=10)
_THIN = Side(style="thin", color="D1D5DB")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_MONEY = 'R$ #,##0.00'


def _write_header(ws, headers: list[str], row: int = 1) -> None:
    for col, label in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=label)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _BORDER
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def _autosize(ws, min_width: int = 10, max_width: int = 52) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        longest = max((len(str(c.value)) for c in col if c.value is not None), default=0)
        ws.column_dimensions[letter].width = max(min_width, min(max_width, longest + 3))


def _period_label(date_start: str | None, date_end: str | None) -> str:
    if not date_start or not date_end:
        return "Todo o período"
    try:
        s = datetime.strptime(date_start[:10], "%Y-%m-%d").strftime("%d/%m/%Y")
        e = datetime.strptime(date_end[:10], "%Y-%m-%d").strftime("%d/%m/%Y")
    except ValueError:
        return f"{date_start} a {date_end}"
    return f"{s} a {e}" if s != e else s


async def build_performance_workbook(
    date_start: str | None = None,
    date_end: str | None = None,
    consultant_email: str | None = None,
    period_label: str | None = None,
    user: dict = None,
) -> tuple[bytes, str]:
    """
    Returns (xlsx bytes, suggested filename).
    """
    if date_start and date_end:
        # Raises ValueError on a bad window; the router maps it to a 400.
        agenda_service.parse_date_range(date_start, date_end)

    kanban = await negocios_service.get_kanban_stats(
        date_start=date_start,
        date_end=date_end,
        consultant_email=consultant_email,
        user=user,
    )
    history = await negocios_service.get_negocios_historico(
        user=user, date_start=date_start, date_end=date_end
    )
    consultants = await leads_service.get_consultants_performance()

    # get_kanban_stats and get_negocios_historico already narrow a consultor to their own data.
    # The consultants sheet must follow the same rule, otherwise the workbook would mix one
    # person's pipeline with everyone else's performance.
    if user and user.get("role") == "consultor":
        consultants = [c for c in consultants if c.get("email") == user.get("email")]

    # The operator filter is an e-mail everywhere except the agenda queries, which match on the
    # recorded name.
    operator_name = None
    if consultant_email:
        history = [h for h in history if h.get("usuario_email") == consultant_email]
        consultants = [c for c in consultants if c.get("email") == consultant_email]
        operator_name = next((c.get("consultant") for c in consultants), None) or next(
            (h.get("usuario_nome") for h in history), None
        )

    agenda = None
    if date_start and date_end:
        try:
            agenda = await agenda_service.get_agenda_performance(
                date_start, date_end, operator_name or "all"
            )
        except ValueError:
            agenda = None

    wb = Workbook()

    # ---------------------------------------------------------------- Resumo
    ws = wb.active
    ws.title = "Resumo"
    ws["A1"] = "Performance Comercial"
    ws["A1"].font = _TITLE_FONT

    meta = [
        ("Período", period_label or _period_label(date_start, date_end)),
        ("Operador", operator_name or consultant_email or "Todos os operadores"),
        ("Gerado em", datetime.now().strftime("%d/%m/%Y %H:%M")),
    ]
    row = 3
    for label, value in meta:
        ws.cell(row=row, column=1, value=label).font = _LABEL_FONT
        ws.cell(row=row, column=2, value=value)
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Pipeline (retrato atual)").font = _TITLE_FONT
    row += 1
    ws.cell(row=row, column=1, value="Não muda com o filtro de período.").font = Font(
        italic=True, size=9, color="6B7280"
    )
    row += 1

    s = kanban["summary"]
    pipeline_rows = [
        ("Total de negócios", s["total_deals"], None),
        ("Em andamento", s["em_andamento"], None),
        ("Ganhos", s["ganhos"], None),
        ("Valor dos ganhos", s["ganhos_valor"], _MONEY),
        ("Perdidos", s["perdidos"], None),
        ("Taxa de ganho (%)", s["win_rate"], '0.0'),
        ("Valor total do pipeline", s["total_valor"], _MONEY),
    ]
    for label, value, fmt in pipeline_rows:
        ws.cell(row=row, column=1, value=label).font = _LABEL_FONT
        cell = ws.cell(row=row, column=2, value=value)
        if fmt:
            cell.number_format = fmt
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Movimentações no período").font = _TITLE_FONT
    row += 1
    total_valor_movido = sum((h.get("valor") or 0) for h in history)
    for label, value, fmt in [
        ("Total de movimentações", len(history), None),
        ("Volume movimentado", total_valor_movido, _MONEY),
    ]:
        ws.cell(row=row, column=1, value=label).font = _LABEL_FONT
        cell = ws.cell(row=row, column=2, value=value)
        if fmt:
            cell.number_format = fmt
        row += 1

    if agenda:
        row += 1
        ws.cell(row=row, column=1, value="Agenda no período").font = _TITLE_FONT
        row += 1
        a = agenda["summary"]
        for label, value, fmt in [
            ("Agendamentos", a["total"], None),
            ("Concluídos", a["completed"], None),
            ("Pendentes", a["pending"], None),
            ("Taxa de conclusão (%)", a["completion_rate"], '0.0'),
        ]:
            ws.cell(row=row, column=1, value=label).font = _LABEL_FONT
            cell = ws.cell(row=row, column=2, value=value)
            if fmt:
                cell.number_format = fmt
            row += 1

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 30

    # ------------------------------------------------------- Etapas do Kanban
    ws = wb.create_sheet("Etapas do Kanban")
    _write_header(ws, ["Etapa", "Atual", "Distribuição (%)", "Valor", "Entradas no período", "Leads distintos"])
    for i, stage in enumerate(kanban["stages"], start=2):
        ws.cell(row=i, column=1, value=stage["etapa"] + (" (fora do padrão)" if stage.get("unknown_stage") else ""))
        ws.cell(row=i, column=2, value=stage["current"])
        ws.cell(row=i, column=3, value=stage["share"]).number_format = '0.0'
        ws.cell(row=i, column=4, value=stage["valor"]).number_format = _MONEY
        ws.cell(row=i, column=5, value=stage["entered"])
        ws.cell(row=i, column=6, value=stage["entered_leads"])
    _autosize(ws)

    # ---------------------------------------------------------- Consultores
    ws = wb.create_sheet("Consultores")
    _write_header(ws, ["Consultor", "E-mail", "Total de leads", "Agendados", "Em follow-up", "Conversão (%)"])
    for i, c in enumerate(consultants, start=2):
        ws.cell(row=i, column=1, value=c.get("consultant"))
        ws.cell(row=i, column=2, value=c.get("email"))
        ws.cell(row=i, column=3, value=c.get("total_leads"))
        ws.cell(row=i, column=4, value=c.get("leads_agendados"))
        ws.cell(row=i, column=5, value=c.get("leads_follow_up"))
        ws.cell(row=i, column=6, value=c.get("conversion_rate")).number_format = '0.00'
    _autosize(ws)

    # ------------------------------------------------------- Agenda por dia
    if agenda:
        ws = wb.create_sheet("Agenda por Dia")
        _write_header(ws, ["Data", "Agendamentos", "Concluídos", "Pendentes", "Conclusão (%)"])
        for i, d in enumerate(agenda["daily"], start=2):
            try:
                shown = datetime.strptime(d["date"], "%Y-%m-%d").strftime("%d/%m/%Y")
            except ValueError:
                shown = d["date"]
            ws.cell(row=i, column=1, value=shown)
            ws.cell(row=i, column=2, value=d["total"])
            ws.cell(row=i, column=3, value=d["completed"])
            ws.cell(row=i, column=4, value=d["pending"])
            rate = (d["completed"] / d["total"] * 100) if d["total"] else 0
            ws.cell(row=i, column=5, value=round(rate, 1)).number_format = '0.0'
        _autosize(ws)

    # -------------------------------------------------- Histórico detalhado
    ws = wb.create_sheet("Histórico Detalhado")
    _write_header(ws, ["Data/Hora", "Lead", "Operador", "E-mail", "Etapa anterior", "Etapa nova", "Valor"])
    for i, h in enumerate(history, start=2):
        raw = h.get("data_hora") or ""
        # Timestamps arrive either ISO ("T") or space separated, depending on the source table.
        shown = raw
        for fmt in ("%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S"):
            try:
                shown = datetime.strptime(raw[:19], fmt).strftime("%d/%m/%Y %H:%M")
                break
            except ValueError:
                continue
        ws.cell(row=i, column=1, value=shown)
        ws.cell(row=i, column=2, value=h.get("lead_name"))
        ws.cell(row=i, column=3, value=h.get("usuario_nome"))
        ws.cell(row=i, column=4, value=h.get("usuario_email"))
        ws.cell(row=i, column=5, value=h.get("etapa_anterior"))
        ws.cell(row=i, column=6, value=h.get("etapa_nova"))
        ws.cell(row=i, column=7, value=h.get("valor") or 0).number_format = _MONEY
    _autosize(ws)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    stamp = datetime.now().strftime("%Y-%m-%d_%H%M")
    filename = f"performance_comercial_{stamp}.xlsx"
    return buffer.getvalue(), filename
